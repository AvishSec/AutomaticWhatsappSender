import sys
import pyautogui
from openpyxl import load_workbook
import time
import os

wb = load_workbook("Kitap1.xlsx")
ws = wb.active

screenWidth, screenHeight = pyautogui.size()
currentMouseX, currentMouseY = pyautogui.position()

resim = os.path.join("ımg", "button1.PNG")

isimler = []
tel = []
mesaj = []


print("Göndermek istediğin kişilerin ismini biliyorsan virgül(,) bırakarak yazınız (boşluk bırakmayın) ")
print("Eğer sadece tek kişiye mesaj atmak istiyorsanız devam edin değilse (h) tuşuna basınız")
search = input("Tel no veya Grup yazınız:")

"""
-----------------YYYYYYYYYYYYYYYY---------------
*
*
*
X
X
X
X
X
*
*
*

sol yukarı -- 697 957
sol aşağı  -- 524 950

1920 1080
mean -- 734 812  2,5==5/2  1,2==6/5

sağ yukarı -- 665 863
sap aşağı -- 640 911

"""

line = 3
col = 5

def telcall():#düzenle
    for i in range(len(tel)):
        print(tel[i])

def excel(col):
    try:
        print(ws.max_row,ws.max_column)
        for x in range(2, ws.max_row+1): #6 = SATIR + 1
            isimler.append(ws.cell(x, 1).value)
            tel.append(ws.cell(x, 2).value)
            mesaj.append(ws.cell(x, 3).value)

            print(isimler[x-2], tel[x-2],mesaj[x-2])

            #print(ws.cell(line, col).value)  2.satır 1.sütun

    except ValueError:
        pass

def calculatr():
    try:
        time.sleep(6)
        #pyautogui.moveTo(1, 10)
        pyautogui.click(resim)

    except TypeError:
        print("Başlattınktan sonra ekranda whatsapp web açık olmalıdır")
        print("Lütfen Whatsapp web de hiçbir yere tıklamayınız")
        sys.exit()

    if search == "h" or search == "H":
        loop = True
        while loop:
            try:
                i = 0
                pyautogui.write(isimler[i],interval=0.25)
                pyautogui.click(os.path.join("ımg","sohbet.PNG"))
                # or pyautogui.press(enter)
                pyautogui.click(os.path.join("ımg","mess.PNG"))
                # or pyautogui.press(enter)
                pyautogui.write("Ben bot sevimli bot",mesaj[0],interval=0.50)
                pyautogui.press("enter")
                pyautogui.click(resim)
                i+=1
                if ws.max_row == (i-1):
                    loop = False

            except TypeError:
                print("Başlattınktan sonra ekranda whatsapp web açık olmalıdır")
                print("Lütfen Whatsapp web de hiçbir yere tıklamayınız")
                sys.exit()
        sys.exit()

    else:
        searchsplit = search.split(",")
        for i in range(len(searchsplit)):
            print(searchsplit[i])
            pyautogui.write(searchsplit[i], interval=0.25)
            pyautogui.press("enter")
            pyautogui.write(mesaj[0], interval=0.20)
            pyautogui.press("enter")
            pyautogui.click(resim)
            print(screenWidth, screenHeight, currentMouseX, currentMouseY)
        sys.exit()


def main():
    excel(col)
    #calculatr()


if __name__ == '__main__':
    main()